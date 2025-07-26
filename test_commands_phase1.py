#!/usr/bin/env python3
"""
Phase 1 Individual Component Tests
Quick command-line tests for specific functionality
"""

import sys
from pathlib import Path
import pandas as pd
import openpyxl

# Add src directory to path
sys.path.append(str(Path(__file__).parent / "src"))

def test_pandas_read():
    """Test 1: Basic pandas reading"""
    print("🧪 Test 1: Reading Excel with pandas...")
    
    try:
        # Bank file
        bank_df = pd.read_excel("data/sample/KH_Bank.XLSX", engine='openpyxl')
        print(f"✅ Bank file: {bank_df.shape[0]} rows, {bank_df.shape[1]} columns")
        
        # Ledger file
        ledger_df = pd.read_excel("data/sample/Customer_Ledger_Entries_FULL.xlsx", engine='openpyxl')
        print(f"✅ Ledger file: {ledger_df.shape[0]} rows, {ledger_df.shape[1]} columns")
        
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_sheet_info():
    """Test 2: Excel sheet information"""
    print("\n🧪 Test 2: Getting sheet information...")
    
    try:
        # Bank file sheets
        bank_excel = pd.ExcelFile("data/sample/KH_Bank.XLSX", engine='openpyxl')
        print(f"✅ Bank sheets: {bank_excel.sheet_names}")
        
        # Ledger file sheets
        ledger_excel = pd.ExcelFile("data/sample/Customer_Ledger_Entries_FULL.xlsx", engine='openpyxl')
        print(f"✅ Ledger sheets: {ledger_excel.sheet_names}")
        
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_column_names():
    """Test 3: Extract column names"""
    print("\n🧪 Test 3: Extracting column names...")
    
    try:
        # Bank columns
        bank_df = pd.read_excel("data/sample/KH_Bank.XLSX", engine='openpyxl')
        print(f"✅ Bank columns ({len(bank_df.columns)}): {list(bank_df.columns[:3])}...")
        
        # Ledger columns
        ledger_df = pd.read_excel("data/sample/Customer_Ledger_Entries_FULL.xlsx", engine='openpyxl')
        print(f"✅ Ledger columns ({len(ledger_df.columns)}): {list(ledger_df.columns[:3])}...")
        
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_data_preview():
    """Test 4: Data preview"""
    print("\n🧪 Test 4: Data preview...")
    
    try:
        # Bank preview
        bank_df = pd.read_excel("data/sample/KH_Bank.XLSX", engine='openpyxl', nrows=3)
        print(f"✅ Bank data preview (3 rows): {bank_df.shape}")
        
        # Ledger preview
        ledger_df = pd.read_excel("data/sample/Customer_Ledger_Entries_FULL.xlsx", engine='openpyxl', nrows=3)
        print(f"✅ Ledger data preview (3 rows): {ledger_df.shape}")
        
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_openpyxl_direct():
    """Test 5: Direct openpyxl usage"""
    print("\n🧪 Test 5: Direct openpyxl access...")
    
    try:
        # Bank file
        bank_wb = openpyxl.load_workbook("data/sample/KH_Bank.XLSX", read_only=True)
        print(f"✅ Bank workbook sheets: {bank_wb.sheetnames}")
        bank_wb.close()
        
        # Ledger file
        ledger_wb = openpyxl.load_workbook("data/sample/Customer_Ledger_Entries_FULL.xlsx", read_only=True)
        print(f"✅ Ledger workbook sheets: {ledger_wb.sheetnames}")
        ledger_wb.close()
        
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_memory_usage():
    """Test 6: Memory usage assessment"""
    print("\n🧪 Test 6: Memory usage assessment...")
    
    try:
        # Bank file memory
        bank_df = pd.read_excel("data/sample/KH_Bank.XLSX", engine='openpyxl')
        bank_memory = bank_df.memory_usage(deep=True).sum() / 1024 / 1024
        print(f"✅ Bank file memory usage: {bank_memory:.2f} MB")
        
        # Ledger file memory
        ledger_df = pd.read_excel("data/sample/Customer_Ledger_Entries_FULL.xlsx", engine='openpyxl')
        ledger_memory = ledger_df.memory_usage(deep=True).sum() / 1024 / 1024
        print(f"✅ Ledger file memory usage: {ledger_memory:.2f} MB")
        
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == "__main__":
    print("🚀 PHASE 1: INDIVIDUAL COMPONENT TESTS")
    print("=" * 50)
    
    tests = [
        test_pandas_read,
        test_sheet_info,
        test_column_names,
        test_data_preview,
        test_openpyxl_direct,
        test_memory_usage
    ]
    
    passed = 0
    total = len(tests)
    
    for test in tests:
        if test():
            passed += 1
    
    print(f"\n📊 RESULTS: {passed}/{total} tests passed")
    
    if passed == total:
        print("🎉 ALL COMPONENT TESTS PASSED!")
    else:
        print("⚠️ Some tests failed - check the errors above")
