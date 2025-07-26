#!/usr/bin/env python3
"""
Phase 1 Testing: Basic Excel Processing
Tests the fundamental Excel reading and processing capabilities
"""

import sys
from pathlib import Path
import pandas as pd
import openpyxl

# Add src directory to path
sys.path.append(str(Path(__file__).parent / "src"))

from financial_analyzer import FinancialDataAnalyzer
from rich.console import Console
from rich.panel import Panel
from rich.text import Text
from rich.table import Table

def test_basic_excel_reading():
    """Test basic Excel file reading with pandas and openpyxl"""
    console = Console()
    
    console.print(Panel(Text("🧪 PHASE 1: BASIC EXCEL PROCESSING TEST", style="bold cyan"), expand=False))
    
    # Define file paths
    bank_file = Path("data/sample/KH_Bank.XLSX")
    ledger_file = Path("data/sample/Customer_Ledger_Entries_FULL.xlsx")
    
    console.print(f"\n[yellow]📂 Testing file access...[/yellow]")
    
    # Check file existence
    if not bank_file.exists():
        console.print(f"[red]❌ Bank file not found: {bank_file}[/red]")
        return False
    console.print(f"[green]✅ Bank file found: {bank_file}[/green]")
    
    if not ledger_file.exists():
        console.print(f"[red]❌ Ledger file not found: {ledger_file}[/red]")
        return False
    console.print(f"[green]✅ Ledger file found: {ledger_file}[/green]")
    
    # Test 1: Basic pandas reading
    console.print(f"\n[yellow]🔍 Test 1: Basic pandas Excel reading...[/yellow]")
    try:
        # Read bank file
        bank_excel = pd.ExcelFile(str(bank_file), engine='openpyxl')
        console.print(f"[green]✅ Successfully opened bank file with pandas[/green]")
        console.print(f"   📊 Sheets found: {bank_excel.sheet_names}")
        
        # Read ledger file  
        ledger_excel = pd.ExcelFile(str(ledger_file), engine='openpyxl')
        console.print(f"[green]✅ Successfully opened ledger file with pandas[/green]")
        console.print(f"   📊 Sheets found: {ledger_excel.sheet_names}")
        
    except Exception as e:
        console.print(f"[red]❌ Error reading with pandas: {e}[/red]")
        return False
    
    # Test 2: openpyxl reading
    console.print(f"\n[yellow]🔍 Test 2: openpyxl reading...[/yellow]")
    try:
        # Open with openpyxl
        bank_wb = openpyxl.load_workbook(str(bank_file), read_only=True)
        console.print(f"[green]✅ Successfully opened bank file with openpyxl[/green]")
        console.print(f"   📊 Worksheets: {bank_wb.sheetnames}")
        
        ledger_wb = openpyxl.load_workbook(str(ledger_file), read_only=True)
        console.print(f"[green]✅ Successfully opened ledger file with openpyxl[/green]")
        console.print(f"   📊 Worksheets: {ledger_wb.sheetnames}")
        
        # Close workbooks
        bank_wb.close()
        ledger_wb.close()
        
    except Exception as e:
        console.print(f"[red]❌ Error reading with openpyxl: {e}[/red]")
        return False
    
    # Test 3: Multiple worksheet handling
    console.print(f"\n[yellow]🔍 Test 3: Multiple worksheet processing...[/yellow]")
    try:
        # Process each sheet in bank file
        for sheet_name in bank_excel.sheet_names:
            df = bank_excel.parse(sheet_name)
            console.print(f"   📄 Bank Sheet '{sheet_name}': {df.shape[0]} rows × {df.shape[1]} columns")
        
        # Process each sheet in ledger file
        for sheet_name in ledger_excel.sheet_names:
            df = ledger_excel.parse(sheet_name)
            console.print(f"   📄 Ledger Sheet '{sheet_name}': {df.shape[0]} rows × {df.shape[1]} columns")
            
    except Exception as e:
        console.print(f"[red]❌ Error processing worksheets: {e}[/red]")
        return False
    
    # Test 4: Display detailed file information
    console.print(f"\n[yellow]🔍 Test 4: Detailed file information extraction...[/yellow]")
    
    # Bank file details
    bank_sheet = bank_excel.parse(bank_excel.sheet_names[0])
    bank_table = Table(title="Bank File Information (KH_Bank.XLSX)")
    bank_table.add_column("Property", style="cyan")
    bank_table.add_column("Value", style="white")
    
    bank_table.add_row("File Size", f"{bank_file.stat().st_size:,} bytes")
    bank_table.add_row("Number of Sheets", str(len(bank_excel.sheet_names)))
    bank_table.add_row("Main Sheet Name", bank_excel.sheet_names[0])
    bank_table.add_row("Dimensions", f"{bank_sheet.shape[0]:,} rows × {bank_sheet.shape[1]} columns")
    bank_table.add_row("Data Types", f"{len(bank_sheet.dtypes.unique())} unique types")
    bank_table.add_row("Memory Usage", f"{bank_sheet.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB")
    
    console.print(bank_table)
    
    # Ledger file details
    ledger_sheet = ledger_excel.parse(ledger_excel.sheet_names[0])
    ledger_table = Table(title="Ledger File Information (Customer_Ledger_Entries_FULL.xlsx)")
    ledger_table.add_column("Property", style="cyan")
    ledger_table.add_column("Value", style="white")
    
    ledger_table.add_row("File Size", f"{ledger_file.stat().st_size:,} bytes")
    ledger_table.add_row("Number of Sheets", str(len(ledger_excel.sheet_names)))
    ledger_table.add_row("Main Sheet Name", ledger_excel.sheet_names[0])
    ledger_table.add_row("Dimensions", f"{ledger_sheet.shape[0]:,} rows × {ledger_sheet.shape[1]} columns")
    ledger_table.add_row("Data Types", f"{len(ledger_sheet.dtypes.unique())} unique types")
    ledger_table.add_row("Memory Usage", f"{ledger_sheet.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB")
    
    console.print(ledger_table)
    
    # Test 5: Column name analysis
    console.print(f"\n[yellow]🔍 Test 5: Column name analysis...[/yellow]")
    
    console.print(f"\n[blue]Bank File Column Names (first 10):[/blue]")
    for i, col in enumerate(bank_sheet.columns[:10]):
        console.print(f"  {i+1:2d}. {col}")
    if len(bank_sheet.columns) > 10:
        console.print(f"     ... and {len(bank_sheet.columns) - 10} more columns")
    
    console.print(f"\n[blue]Ledger File Column Names (first 10):[/blue]")
    for i, col in enumerate(ledger_sheet.columns[:10]):
        console.print(f"  {i+1:2d}. {col}")
    if len(ledger_sheet.columns) > 10:
        console.print(f"     ... and {len(ledger_sheet.columns) - 10} more columns")
    
    # Test 6: Data quality check
    console.print(f"\n[yellow]🔍 Test 6: Basic data quality assessment...[/yellow]")
    
    # Bank data quality
    bank_null_count = bank_sheet.isnull().sum().sum()
    bank_total_cells = bank_sheet.shape[0] * bank_sheet.shape[1]
    bank_null_percentage = (bank_null_count / bank_total_cells) * 100
    
    console.print(f"[blue]Bank Data Quality:[/blue]")
    console.print(f"  • Total cells: {bank_total_cells:,}")
    console.print(f"  • Null cells: {bank_null_count:,} ({bank_null_percentage:.1f}%)")
    console.print(f"  • Non-null cells: {bank_total_cells - bank_null_count:,}")
    
    # Ledger data quality
    ledger_null_count = ledger_sheet.isnull().sum().sum()
    ledger_total_cells = ledger_sheet.shape[0] * ledger_sheet.shape[1]
    ledger_null_percentage = (ledger_null_count / ledger_total_cells) * 100
    
    console.print(f"[blue]Ledger Data Quality:[/blue]")
    console.print(f"  • Total cells: {ledger_total_cells:,}")
    console.print(f"  • Null cells: {ledger_null_count:,} ({ledger_null_percentage:.1f}%)")
    console.print(f"  • Non-null cells: {ledger_total_cells - ledger_null_count:,}")
    
    return True

def test_financial_analyzer_integration():
    """Test integration with FinancialDataAnalyzer class"""
    console = Console()
    
    console.print(f"\n[yellow]🔍 Test 7: FinancialDataAnalyzer integration...[/yellow]")
    
    try:
        analyzer = FinancialDataAnalyzer()
        console.print(f"[green]✅ FinancialDataAnalyzer initialized successfully[/green]")
        
        # Test bank statement loading
        bank_info = analyzer.load_bank_statement("data/sample/KH_Bank.XLSX")
        if bank_info:
            console.print(f"[green]✅ Bank statement loaded via analyzer[/green]")
            console.print(f"   📊 Found {len(bank_info)} sheets")
        else:
            console.print(f"[yellow]⚠️ Bank statement loading returned empty[/yellow]")
        
        # Test ledger loading
        ledger_info = analyzer.load_customer_ledger("data/sample/Customer_Ledger_Entries_FULL.xlsx")
        if ledger_info:
            console.print(f"[green]✅ Customer ledger loaded via analyzer[/green]")
            console.print(f"   📊 Found {len(ledger_info)} sheets")
        else:
            console.print(f"[yellow]⚠️ Customer ledger loading returned empty[/yellow]")
        
        return True
        
    except Exception as e:
        console.print(f"[red]❌ Error in analyzer integration: {e}[/red]")
        return False

def main():
    console = Console()
    
    console.print("[bold green]🚀 STARTING PHASE 1 TESTS[/bold green]")
    console.print("="*60)
    
    # Run basic Excel tests
    basic_test_passed = test_basic_excel_reading()
    
    # Run analyzer integration test
    integration_test_passed = test_financial_analyzer_integration()
    
    # Summary
    console.print("\n" + "="*60)
    console.print("[bold yellow]📋 PHASE 1 TEST SUMMARY[/bold yellow]")
    console.print("="*60)
    
    if basic_test_passed:
        console.print("[green]✅ Basic Excel Processing: PASSED[/green]")
    else:
        console.print("[red]❌ Basic Excel Processing: FAILED[/red]")
    
    if integration_test_passed:
        console.print("[green]✅ Analyzer Integration: PASSED[/green]")
    else:
        console.print("[red]❌ Analyzer Integration: FAILED[/red]")
    
    if basic_test_passed and integration_test_passed:
        console.print("\n[bold green]🎉 PHASE 1: ALL TESTS PASSED! 🎉[/bold green]")
        console.print("Your Excel processing foundation is working correctly.")
        console.print("Ready to proceed to Phase 2: Transaction Parsing")
    else:
        console.print("\n[bold red]⚠️ PHASE 1: SOME TESTS FAILED[/bold red]")
        console.print("Please resolve the issues before proceeding to the next phase.")
    
    console.print("\n[cyan]Files successfully processed:[/cyan]")
    console.print("  📄 KH_Bank.XLSX - Bank statement with complex structure")
    console.print("  📄 Customer_Ledger_Entries_FULL.xlsx - Customer ledger entries")

if __name__ == "__main__":
    main()
